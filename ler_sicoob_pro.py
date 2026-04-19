import streamlit as st
import fitz
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import warnings
import time

warnings.simplefilter(action='ignore', category=FutureWarning)

st.set_page_config(page_title="Consolidador SICOOB", page_icon="📊", layout="wide")

# ==========================================
# INJEÇÃO DE CSS: CARDS BONITOS E CORES SICOOB
# ==========================================
st.markdown("""
<style>
    /* Deixa o fundo do site levemente cinza para destacar os cards brancos */
    .stApp {
        background-color: #f8fafc;
    }
    
    /* Estilização dos Cards das Métricas */
    [data-testid="stMetric"] {
        background-color: #ffffff;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.05);
        border-left: 6px solid #00ae9d; /* Ciano SICOOB */
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    
    /* Efeito ao passar o mouse por cima do Card */
    [data-testid="stMetric"]:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 15px rgba(0, 0, 0, 0.1);
    }

    /* Título do Card */
    [data-testid="stMetricLabel"] {
        font-size: 16px !important;
        font-weight: 600 !important;
        color: #475569 !important;
    }

    /* Valor Numérico do Card */
    [data-testid="stMetricValue"] {
        font-size: 32px !important;
        font-weight: 800 !important;
        color: #003641 !important; /* Verde Escuro SICOOB */
    }
    
    /* Esconder o cabeçalho padrão do Streamlit (opcional, deixa mais clean) */
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# Cabeçalho da Aplicação
col_logo, col_titulo = st.columns([1, 8])
with col_logo:
    st.markdown("<h1 style='text-align: center; color: #00ae9d;'>📊</h1>", unsafe_allow_html=True)
with col_titulo:
    st.markdown("<h1 style='color: #003641; margin-bottom: 0px;'>Consolidador SICOOB Pro</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #64748b;'>Extração, consolidação e formatação automática de extratos em PDF.</p>", unsafe_allow_html=True)

st.divider()

# Criação de Abas para Organizar a Tela
aba_processamento, aba_auditoria = st.tabs(["📤 Upload e Resumo", "🔎 Auditoria de Dados"])

with aba_processamento:
    arquivos_pdf = st.file_uploader("Arraste os seus extratos em PDF para aqui", type="pdf", accept_multiple_files=True)

    if arquivos_pdf:
        if st.button("Processar Documentos", type="primary"):
            
            # Barra de progresso para UX
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            dados_totais = []
            total_arquivos = len(arquivos_pdf)
            
            for i, arquivo in enumerate(arquivos_pdf):
                status_text.text(f"Analisando arquivo {i+1} de {total_arquivos}: {arquivo.name}...")
                try:
                    doc = fitz.open(stream=arquivo.read(), filetype="pdf")
                    for page in doc:
                        for t in page.find_tables():
                            df = t.to_pandas()
                            if any('Sacado' in col for col in df.columns) and any('Valor (R$)' in col for col in df.columns):
                                df['Arquivo_Origem'] = arquivo.name 
                                dados_totais.append(df)
                    doc.close()
                except Exception as e:
                    st.warning(f"Não foi possível ler o arquivo {arquivo.name}. Ele foi ignorado.")
                
                # Atualiza a barra de progresso
                progress_bar.progress((i + 1) / total_arquivos)

            status_text.text("Estruturando dados...")
            time.sleep(0.5) # Pausa dramática para o usuário ver que terminou
            status_text.empty()
            progress_bar.empty()

            if not dados_totais:
                st.error("Nenhuma tabela válida encontrada nos PDFs enviados! Verifique se são relatórios do SICOOB.")
            else:
                df_raw = pd.concat(dados_totais, ignore_index=True)

                cols_esperadas = ['Sacado','Nosso Número','Seu Número','Dt. Previsão Crédito',
                                  'Vencimento','Dt. Limite\nPgto','Valor (R$)','Vlr. Mora',
                                  'Vlr. Desc.','Vlr. Outros\nAcresc.','Dt. Liquid.','Vlr. Cobrado', 'Arquivo_Origem']
                cols_presentes = [c for c in cols_esperadas if c in df_raw.columns]
                df = df_raw[cols_presentes].copy()

                mapa_nomes = {
                    'Sacado': 'Sacado', 'Nosso Número': 'Nosso_Numero', 'Seu Número': 'Seu_Numero',
                    'Dt. Previsão Crédito': 'Prev_Credito', 'Vencimento': 'Vencimento',
                    'Dt. Limite\nPgto': 'Dt_Limite', 'Valor (R$)': 'Valor_Original', 'Vlr. Mora': 'Mora',
                    'Vlr. Desc.': 'Desconto', 'Vlr. Outros\nAcresc.': 'Outros',
                    'Dt. Liquid.': 'Dt_Liquidacao', 'Vlr. Cobrado': 'Valor_Cobrado', 'Arquivo_Origem': 'Origem'
                }
                df.rename(columns=mapa_nomes, inplace=True)

                if 'Sacado' in df.columns:
                    df['Sacado'] = df['Sacado'].astype(str).str.replace('\n',' ', regex=False).str.strip()
                    df['Sacado'] = df['Sacado'].str.replace(r'\d{11}|\d{14}','', regex=True).str.strip()

                colunas_moeda = ['Valor_Original','Mora','Desconto','Outros','Valor_Cobrado']
                for col in colunas_moeda:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.replace('.','', regex=False).str.replace(',','.', regex=False)
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).round(2)

                colunas_data = ['Prev_Credito','Vencimento','Dt_Limite','Dt_Liquidacao']
                for col in colunas_data:
                    if col in df.columns:
                        df[col] = pd.to_datetime(df[col], format='%d/%m/%Y', errors='coerce')

                df = df.replace({pd.NaT: None, float('nan'): None})
                
                # Remoção de Linhas Vazias e Duplicadas (Proteção)
                if 'Nosso_Numero' in df.columns:
                    linhas_antes = len(df)
                    df = df.dropna(subset=['Nosso_Numero'])
                    df = df.drop_duplicates(subset=['Nosso_Numero'])
                    linhas_depois = len(df)
                    if linhas_antes != linhas_depois:
                        st.info(f"Limpeza Automática: {linhas_antes - linhas_depois} títulos duplicados foram removidos da consolidação.")

                # ==========================================
                # INDICADORES (CARDS CSS APLICADOS AQUI)
                # ==========================================
                st.subheader("Resumo da Consolidação")
                col1, col2, col3 = st.columns(3)
                
                total_titulos = len(df)
                valor_total = df['Valor_Cobrado'].sum() if 'Valor_Cobrado' in df.columns else 0
                mora_total = df['Mora'].sum() if 'Mora' in df.columns else 0
                
                str_valor = f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                str_mora = f"R$ {mora_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                col1.metric(label="Títulos Processados", value=total_titulos)
                col2.metric(label="Total Liquidado", value=str_valor)
                col3.metric(label="Total de Mora", value=str_mora)

                # Gerar Excel em Memória
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Dados', index=False, startrow=4)
                    wb = writer.book
                    ws = wb['Dados']
                    
                    ws.sheet_view.showGridLines = False
                    
                    ws['A1'] = 'RELATÓRIO DE LIQUIDAÇÕES - SICOOB'
                    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='404040')
                    ws['A2'] = f'Fontes: {len(arquivos_pdf)} ficheiro(s) PDF processado(s)'
                    ws['A2'].font = Font(name='Calibri', size=11, color='595959')
                    ws['A3'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
                    ws['A3'].font = Font(size=9, italic=True, color='808080')
                    
                    header_row = 5
                    last_data_row = header_row + len(df)
                    last_col = get_column_letter(len(df.columns))
                    
                    tabela = Table(displayName="TbSicoob", ref=f"A{header_row}:{last_col}{last_data_row}")
                    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True, showColumnStripes=False)
                    ws.add_table(tabela)
                    ws.freeze_panes = 'A6'
                    
                    formato_moeda = '#,##0.00' 
                    formato_data = 'DD/MM/YYYY'
                    
                    for idx, nome_col in enumerate(df.columns, 1):
                        letra = get_column_letter(idx)
                        if nome_col == 'Sacado': ws.column_dimensions[letra].width = 40
                        elif nome_col == 'Origem': ws.column_dimensions[letra].width = 25
                        elif nome_col in ['Nosso_Numero','Seu_Numero']: ws.column_dimensions[letra].width = 16
                        else: ws.column_dimensions[letra].width = 14
                        
                        for row in range(header_row+1, last_data_row+1):
                            cell = ws[f'{letra}{row}']
                            if nome_col in colunas_moeda: cell.number_format = formato_moeda
                            elif nome_col in colunas_data:
                                if cell.value is not None: cell.number_format = formato_data
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else: cell.alignment = Alignment(vertical='center')

                    linha_total = last_data_row + 1
                    ws[f'A{linha_total}'] = "TOTAIS (Visíveis):"
                    ws[f'A{linha_total}'].font = Font(name='Calibri', size=11, bold=True, color='404040')
                    ws[f'A{linha_total}'].alignment = Alignment(horizontal='right', vertical='center')
                    
                    col_inicio = df.columns.get_loc('Valor_Original') if 'Valor_Original' in df.columns else 0
                    if col_inicio > 1:
                        ws.merge_cells(f'A{linha_total}:{get_column_letter(col_inicio)}{linha_total}')

                    for idx, nome_col in enumerate(df.columns, 1):
                        if nome_col in colunas_moeda:
                            letra = get_column_letter(idx)
                            cel = ws[f'{letra}{linha_total}']
                            cel.value = f"=SUBTOTAL(109,{letra}{header_row+1}:{letra}{last_data_row})"
                            cel.number_format = formato_moeda
                            cel.font = Font(name='Calibri', size=11, bold=True, color='1F1F1F')
                            cel.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                            cel.border = Border(top=Side(style='thin', color='BFBFBF'), bottom=Side(style='medium', color='BFBFBF'))

                st.write("") # Espaçamento
                st.success("✅ Tudo pronto! O cruzamento de dados foi concluído com sucesso.")
                
                # Centralizando o botão de download para chamar mais atenção
                col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                with col_btn2:
                    st.download_button(
                        label="📥 BAIXAR PLANILHA CONSOLIDADA",
                        data=output.getvalue(),
                        file_name=f"RELATORIO_SICOOB_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                # ==========================================
                # ABA 2: VISUALIZAÇÃO INTERATIVA (DATAFRAME)
                # ==========================================
                with aba_auditoria:
                    st.markdown("### Pré-visualização da Tabela")
                    st.caption("Esta é uma prévia dos dados limpos. Para realizar filtros complexos ou imprimir, baixe a planilha na aba Upload.")
                    
                    df_view = df.copy()
                    for col in colunas_moeda:
                        if col in df_view.columns:
                            df_view[col] = df_view[col].apply(lambda x: f"R$ {x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if pd.notnull(x) else "")
                    
                    for col in colunas_data:
                        if col in df_view.columns:
                            df_view[col] = pd.to_datetime(df_view[col]).dt.strftime('%d/%m/%Y')
                    
                    st.dataframe(df_view, use_container_width=True, hide_index=True)
